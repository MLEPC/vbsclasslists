# Reads excel file, keeps specified columns, drops duplicate records, 
# writes to a file, and formats the header and columns
import pandas as pd
import openpyxl as op
from openpyxl.utils import get_column_letter
#import classlistfunctions as f

# define global variables and variables that may change
working_data = "VBSChildinformation2018.xlsx"
class_list = "VBSChildinformation2018.xlsx"
master_sheet = 'Master Child Info'
start_row = 6
start_head = start_row - 1
num_crews = 17

try:
    df = pd.read_excel(working_data, master_sheet)
except FileNotFoundError:
    # the file does not exist
    print('The file does not exist. Create the child information file and run again.')
    exit(0)
else:
    # the file exists - continue
    pass

# add column for crew assignments
df['Attendance'] = 'M   T   W   Th   F'
df['Child Name'] = df['Child Last Name'] + ', ' + df['Child First Name']
df['Parent Name'] = df['Parent Last Name'] + ', ' + df['Parent First Name']
df.rename(columns={'Age During VBS': 'Age', 'Allergies/Concerns': 'Allergy','Cell Phone': 'Cell','Work Phone': 'Work','Best Contact Number': 'Best Contact'}, inplace=True)

# Keep only columns relevant to class lists
keep_col = ['Crew','Attendance','Child Name','Gender','Age','Allergy','Parent Name','Phone','Work','Cell']
df = df[keep_col]

try:
    xlsx_doc = op.load_workbook(class_list)
except FileNotFoundError:
    # the file does not exist
    # create the workbook and load it
    print('Create the child info file and assign crews before running this program.')
    writer = pd.ExcelWriter(class_list)
    workbook = writer.book
    writer.save()
    xlsx_doc = op.load_workbook(class_list)
    
    for c in range(1,num_crews+1):
        # create dataframe for crew
        crew_df = df.query('Crew == ' + str(c))
        # delete Crew column
        crew_df = crew_df.drop('Crew', axis=1, inplace=False)
        # create a worksheet for list of children
        sheet = xlsx_doc.create_sheet(title = 'Crew ' + str(c))
        sheet = xlsx_doc['Crew ' + str(c)]
        # write the header
        for col_num, value in enumerate(crew_df.columns.values):
            column_letter = get_column_letter(col_num+1)
            sheet[column_letter + str(start_row)] = value
        # write the data to new sheet
        # for given rows and columns, iterate through the cells
        # insert value for dataframe into given cell
        for col in range(1,len(crew_df.columns)+1):
            column_letter = get_column_letter(col)
            for rw in range(start_row+1,len(crew_df)+start_row+1):
                sheet[column_letter + str(rw)] = crew_df.iat[rw-start_row-1,col-1]
    xlsx_doc.save(class_list)
    xlsx_doc.close()
else:
    # the file exists
    for c in range(1,num_crews+1):
        # create dataframe for crew
        crew_df = df.query('Crew == ' + str(c))
        # delete Crew column
        crew_df = crew_df.drop('Crew', axis=1, inplace=False)
        
        # set name of sheet
        crew_sheet = 'Crew ' + str(c)
        # check if sheet exists, create or load
        if crew_sheet in xlsx_doc.sheetnames:
            sheet = xlsx_doc[crew_sheet]
        else:
            sheet = xlsx_doc.create_sheet(title = crew_sheet)
            sheet = xlsx_doc[crew_sheet]
            # write the header
            for col_num, value in enumerate(crew_df.columns.values):
                column_letter = get_column_letter(col_num+1)
                sheet[column_letter + str(start_row)] = value
        
        # clear previous class list
        # account for more than the max number of children so clear 10 rows
        mult_cells = sheet['A7':'K37']
        for row in mult_cells:
            for cell in row:
                cell.value = ''
        # write the data to new sheet
        # for given rows and columns, iterate through the cells
        # insert value for dataframe into given cell
        for col in range(1,len(crew_df.columns)+1):
            column_letter = get_column_letter(col)
            for rw in range(start_row+1,len(crew_df)+start_row+1):
                sheet[column_letter + str(rw)] = crew_df.iat[rw-start_row-1,col-1]
    
    # check if Sheet1 exists, delete Sheet1
    #if 'Sheet1' in xlsx_doc.sheetnames:
    #   xlsx_doc.remove('Sheet1')
    xlsx_doc.save(class_list)
    xlsx_doc.close()
    '''
    # the file exists
    # clear previous class list
    # account for more than the max number of children so clear 10 rows
    sheet = xlsx_doc[working_sheet]
    mult_cells = sheet['A7':'K37']
    for row in mult_cells:
        for cell in row:
            cell.value = ''
    # for given rows and columns, iterate through the cells
    # insert value for dataframe into given cell
    for column in range(1,len(df.columns)+1):
        column_letter = get_column_letter(column)
        for row in range(start_row+1,len(df)+start_row):
            sheet[column_letter + str(row)] = df.iat[row-start_row-1,column-1]
    xlsx_doc.save(class_list)
    xlsx_doc.close()
    '''
    