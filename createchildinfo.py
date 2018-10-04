# Reads excel file, keeps specified columns, drops duplicate records, 
# writes to a file, and formats the header and columns
import pandas as pd
import openpyxl as op
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment
#import win32com.client
#from win32com.client import Dispatch
import childinfofunctions as f
# define variables
new_data = "VBSChildinformation2018_Data.xlsx"
working_data = "VBSChildinformation2018.xlsx"
class_list = "VBSChildInformation2018.xlsx"
sheet_name = 'Worksheet'
master_sheet = 'Master Child Info'

try:
    new_df = pd.read_excel(new_data, sheet_name)
except FileNotFoundError:
    # the file does not exist
    print('The file does not exist. Download the child information file and run again.')
    exit(0)
else:
    # the file exists - continue
    pass

# load lines greater than given Time
# keep only records since the last pull
last_pull = '6/26/2018  7:00:00 AM'
new_df = new_df[(new_df['Time'] > last_pull)]

# add column for crew assignments
new_df['Crew'] = ''

# Keep only columns relevant to class lists
keep_col = ['Crew','Child Last Name','Child First Name','Last Grade Completed','My child requests to be with:','Nickname','Gender','Age During VBS','Tell us about your concern:','Parent First Name','Parent Last Name','Phone','Work Phone','Cell Phone','Best Contact Number']
new_df = new_df[keep_col]
new_df = new_df.rename(index=str, columns={"Tell us about your concern:": "Allergies/Concerns", "My child requests to be with:": "Crewmate Requests"})

# `subset=None` means that every column is used to determine if two rows are 
# different; to change that specify the columns as an array
# `inplace=True` means that the data structure is changed and duplicates 
# are gone 
new_df.drop_duplicates(subset=None, inplace=True)

# convert all names to proper case (initial caps)
# strip leading and trailing spaces
str_to_format = ['Child First Name','Child Last Name','Parent First Name','Parent Last Name','Nickname']
new_df = f.format_str(new_df, str_to_format)

# if working data file does not exist, create it and write new_df to it
# if working data file exists, read contents into working_df, append 
# new_df to working_df, and write combined dataframes to excel output file
try:
    xlsx_doc = op.load_workbook(class_list)
except FileNotFoundError:
    # the file does not exist
    # sort dataframe on last grade completed, child last name, child first name
    new_df = new_df.sort_values(by=['Last Grade Completed','Child Last Name','Child First Name'])
    # create the file and write the results to the file
    writer = pd.ExcelWriter(class_list)
    workbook = writer.book
    new_df.to_excel(writer, master_sheet, index=False, startrow=1, header=False)
    writer = f.format_childinfo(new_df, writer, master_sheet)
    
    writer.save()
else:
    # the file exists, append new_df to file master sheet
    sheet = xlsx_doc[master_sheet]
    row_count=sheet.max_row
    # define formatting
    # format cells text justification and border
    bd = Side(style='thin', color="000000")
    border = Border(left=bd, top=bd, right=bd, bottom=bd)
    align_left = Alignment(vertical="center", horizontal="left")
    align_center = Alignment(vertical="center", horizontal="center")
    # for given rows and columns, iterate through the cells
    # insert value for dataframe into given cell
    for col in range(1,len(new_df.columns)+1):
       column_letter = get_column_letter(col)
       for rw in range(row_count+1,len(new_df)+row_count+1):
           sheet[column_letter + str(rw)] = new_df.iat[rw-row_count-1,col-1]
           sheet[column_letter + str(rw)].border = border
           sheet[column_letter + str(rw)].alignment = align_left
    # check if Sheet1 exists, delete Sheet1
    #if 'Sheet1' in xlsx_doc.sheetnames:
    #   xlsx_doc.remove('Sheet1')
    xlsx_doc.save(class_list)
    xlsx_doc.close()
    
    '''
    # KEEP does not work - troubelshoot
    # sort child information
    excel = win32com.client.Dispatch("Excel.Application")
    wb = excel.Workbooks.Open(class_list)
    ws = wb.Worksheets(master_sheet)
    ws.Range('A1:N' + row_count).Sort(Key1=ws.Range('D1'), Order1=1, Orientation=1)
    wb.Save()   
    excel.Application.Quit()       
    '''
    
    '''
    # KEEP for reference - combine two dataframes
    # the file exists, append new_df to working_df
    working_df = new_df.append(working_df)
    # sort dataframe on crew, child last name, child first name
    working_df = working_df.sort_values(by=['Crew','Child Last Name','Child First Name'])
    # write the results to the file
    writer = pd.ExcelWriter(excel_output)
    working_df.to_excel(writer, sheet_name, index=False, startrow=1, header=False)
    writer = f.format_childinfo(working_df, writer, sheet_name)
    writer.save()
    '''